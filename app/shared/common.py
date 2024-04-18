from abc import ABC, abstractmethod
import os
import random

import numpy as np
from typing import Any, List
import cv2
from PIL import Image
import openpyxl
import keras_cv


class PipelineInterface(ABC):
    def __init__(self):
        self.stages = []

    @abstractmethod
    def add_stage(self, stage):
        pass

    @abstractmethod
    def execute_pipeline(self):
        pass


class Pipeline(PipelineInterface):

    def add_stage(self, stage):
        self.stages.append(stage)

    def execute_pipeline(self):
        for stage in self.stages:
            stage.process()


class FramePreprocessor(ABC):
    """
    This class defines the interface for the frame preprocessing.
    """
    @abstractmethod
    def preprocess_frame(
            self,
            frame: np.ndarray,
            preprocessing_setup: dict[str, Any] | None = None):
        """
        This function preprocesses a frame using OpenCV frame.
        """


class FramePreprocessorOpenCV(FramePreprocessor):
    """
    This class defines the frame preprocessing using OpenCV.
    """

    def preprocess_frame(
            self,
            frame: np.ndarray,
            preprocessing_setup: dict[str, Any] | None = None,
            template_img: np.ndarray | None = None) -> np.ndarray:
        """
        Frame preprocessing.

        Args:
            frame (np.ndarray): Input frame.
            preprocessing_setup (dict): Dictionary containing preprocessing
                parameters. Defaults to None.

        Returns:
            np.ndarray: Preprocessed frame.
        """
        if preprocessing_setup and preprocessing_setup["grayscale"]:
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        if (preprocessing_setup and
                preprocessing_setup["alignment"]["activate"]):
            alignment_dict = preprocessing_setup["alignment"]
            frame = align_image(
                template_img=template_img,
                frame=frame,
                top_match_perc=alignment_dict["top_match_perc"],
                n_features=alignment_dict["n_features"],
                uniform_scaling=alignment_dict["uniform_scaling"])
            # Crop again only if alignment is on
            if (preprocessing_setup and
                    preprocessing_setup["crop"] is not None):
                frame = frame[
                    preprocessing_setup["crop"][0]:
                    preprocessing_setup["crop"][1],
                    preprocessing_setup["crop"][2]:
                    preprocessing_setup["crop"][3]]
        if preprocessing_setup and preprocessing_setup["blur"] is not None:
            kernel_size = (preprocessing_setup["blur"][0],
                           preprocessing_setup["blur"][1])
            sigma = preprocessing_setup["blur"][2]
            frame = cv2.GaussianBlur(frame, kernel_size, sigma)
        if preprocessing_setup and preprocessing_setup["resize"] is not None:
            frame = cv2.resize(frame, preprocessing_setup["resize"])
        if preprocessing_setup and preprocessing_setup["padding"] is not None:
            frame = cv2.copyMakeBorder(
                frame,
                preprocessing_setup["padding"][0],
                preprocessing_setup["padding"][1],
                preprocessing_setup["padding"][2],
                preprocessing_setup["padding"][3],
                cv2.BORDER_CONSTANT,
                value=(128, 128, 128))

        return frame

    def augment_data(
        self, frame: np.ndarray, preprocessing_setup: dict) \
            -> tuple[List[np.ndarray], List[np.ndarray],
                     List[np.ndarray], List[np.ndarray]]:
        """
        Augment input frame using translation, rotation,
        and contrast-brightness adjustments.

        Args:
            frame (np.ndarray): Input frame.
            preprocessing_setup (dict):
            Dictionary containing augmentation parameters.

        Returns:
            tuple: Lists of rotated frames,
            contrast-brightness adjusted frames,
            translated frames, and rototranslated frames.
        """
        contrast_bright_setup = preprocessing_setup[
            "augmentation"]["contrast_brightness"]
        rotation_setup = preprocessing_setup[
            "augmentation"]["rotation"]
        translation_setup = preprocessing_setup[
            "augmentation"]["translation"]
        rototranslation = preprocessing_setup[
            "augmentation"]["rototranslation"]

        translated_frames_list = self.translate(frame, translation_setup)
        rotated_frames_list = self.rotate(frame, rotation_setup)
        if rototranslation:
            rototranslated_frames_list = self.rototranslate(
                frame, translation_setup, rotation_setup)

        tot_frames_list = (translated_frames_list + rotated_frames_list +
                           rototranslated_frames_list)
        tot_frames_list.append(frame)  # Add also the starting frame
        # Choose 'num_start_images' rand numbers in [0, 'len(tot_frames_list)']
        random_numbers = random.sample(
            range(len(tot_frames_list)),
            contrast_bright_setup["num_start_images"])
        selected_frames = [tot_frames_list[number]
                           for number in random_numbers]
        luminance_list = []  # List of frames with luminance transformation
        for img in selected_frames:
            # For each new image apply luminance transformations
            rand_contrast_brightness_list = self.adjust_contrast_brightness(
                img, contrast_bright_setup)
            for cb in rand_contrast_brightness_list:
                luminance_list.append(cb)

        return (rotated_frames_list, translated_frames_list,
                rototranslated_frames_list,
                luminance_list)

    def translate(self, frame: np.ndarray, translation_setup: dict) \
            -> List[np.ndarray]:
        """
        Translate input frame.

        Args:
            frame (np.ndarray): Input frame.
            translation_setup (dict):
            Dictionary containing translation parameters.

        Returns:
            List[np.ndarray]: List of translated frames.
        """
        translated_frames_list = []
        for _ in range(translation_setup["num_translations"]):
            x_trans = random.uniform(translation_setup["x_limit"][0],
                                     translation_setup["x_limit"][1])
            y_trans = random.uniform(translation_setup["y_limit"][0],
                                     translation_setup["y_limit"][1])

            M = np.float32([[1, 0, x_trans], [0, 1, y_trans]])
            height, width = frame.shape[:2]
            translated_frame = cv2.warpAffine(
                frame, M, (width, height), borderMode=cv2.BORDER_DEFAULT)
            translated_frames_list.append(translated_frame)

        return translated_frames_list

    def rotate(
            self, frame: np.ndarray, rotation_setup: dict) -> List[np.ndarray]:
        """
        Rotate input frame.

        Args:
            frame (np.ndarray): Input frame.
            rotation_setup (dict): Dictionary containing rotation parameters.

        Returns:
            List[np.ndarray]: List of rotated frames.
        """
        random_center = random.sample(
            rotation_setup["center_list"], rotation_setup["selected_centers"])
        rotated_frames_list = []

        for center in random_center:
            angles = [random.uniform(rotation_setup["angle_tuple"][0],
                                     rotation_setup["angle_tuple"][1])
                      for _ in range(rotation_setup["selected_angles"])]

            for angle in angles:
                M = cv2.getRotationMatrix2D(
                    center, angle, rotation_setup["scale"])
                height, width = frame.shape[:2]
                rotated_frame = cv2.warpAffine(
                    frame, M, (width, height), borderMode=cv2.BORDER_DEFAULT)
                rotated_frames_list.append(rotated_frame)

        return rotated_frames_list

    def rototranslate(
        self, frame: np.ndarray, translation_setup: dict,
        rotation_setup: dict) \
            -> List[np.ndarray]:
        """
        Rotate and then translate input frame.

        Args:
            frame (np.ndarray): Input frame.
            translation_setup (dict):
            Dictionary containing translation parameters.
            rotation_setup (dict): Dictionary containing rotation parameters.
            rototranslation_setup (dict):
            Dictionary containing rototranslation parameters.

        Returns:
            List[np.ndarray]: List of rototranslated frames.
        """
        rototranslated_frames_list = []
        # Random translation
        translated_frames_list = self.translate(frame, translation_setup)
        # Random rotation
        for translated in translated_frames_list:
            rototranslated_frames_list = self.rotate(translated,
                                                     rotation_setup)

        return rototranslated_frames_list

    def adjust_contrast_brightness(
        self, frame: np.ndarray, contrast_bright_setup: dict) \
            -> List[np.ndarray]:
        """
        Adjust contrast and brightness of input frame.

        Args:
            frame (np.ndarray): Input frame.
            contrast_bright_setup (dict):
            Dictionary containing contrast-brightness adjustment parameters.

        Returns:
            List[np.ndarray]:
            List of frames with adjusted contrast and brightness.
        """
        original_image = frame / 255.0
        original_image = np.expand_dims(original_image, axis=0)
        rand_contrast_brightness_list = []

        for _ in range(contrast_bright_setup["num_transformations"]):
            random_contrast = keras_cv.layers.RandomContrast(
                value_range=(0, 1),
                factor=contrast_bright_setup["contrast_factor"], seed=42)
            transformed_image = random_contrast(original_image)

            random_brightness = keras_cv.layers.RandomBrightness(
                value_range=(0, 1),
                factor=contrast_bright_setup["brightness_factor"], seed=42)
            transformed_image = random_brightness(transformed_image)

            transformed_image = np.squeeze(transformed_image, axis=0)
            transformed_image = (transformed_image * 255).astype(np.uint8)

            rand_contrast_brightness_list.append(transformed_image)

        return rand_contrast_brightness_list


class FramePreprocessorPIL(FramePreprocessor):
    """
    This class defines the frame preprocessing using PIL.
    """

    def preprocess_frame(self,
                         frame: np.ndarray,
                         preprocessing_setup: dict[str, Any] | None = None):
        if preprocessing_setup["crop"] is not None:
            frame = frame[
                preprocessing_setup["crop"][0]:preprocessing_setup["crop"][1],
                preprocessing_setup["crop"][2]:preprocessing_setup["crop"][3]]
        if preprocessing_setup["blur"]:
            # Applicare blurring gaussiano
            kernel_size = (preprocessing_setup["blur"][0],
                           preprocessing_setup["blur"][1])
            # Puoi regolare le dimensioni del kernel a tuo piacimento
            sigma = preprocessing_setup["blur"][2]
            # Puoi regolare il parametro sigma secondo le tue esigenze
            frame = cv2.GaussianBlur(frame, kernel_size, sigma)
        if preprocessing_setup["resize"] is not None:
            frame = cv2.resize(
                frame, preprocessing_setup["resize"])
        if preprocessing_setup["padding"] is not None:
            frame = cv2.copyMakeBorder(
                frame,
                preprocessing_setup["padding"][0],
                preprocessing_setup["padding"][1],
                preprocessing_setup["padding"][2],
                preprocessing_setup["padding"][3],
                cv2.BORDER_CONSTANT,
                value=(128, 128, 128))
        if preprocessing_setup["grayscale"]:
            frame = Image.fromarray(frame).convert("L")

        return np.array(frame)


def get_image_shape(
        preprocessing_setup: dict[str, Any] | None = None) \
        -> tuple[tuple[int, int], int]:
    """
    This function returns the image shape and number of channels
    based on the preprocessing setup.
    """
    if preprocessing_setup['grayscale']:
        number_of_channels = 1
    else:
        number_of_channels = 3
    if preprocessing_setup["resize"] is not None:
        if preprocessing_setup["padding"] is not None:
            if preprocessing_setup['resize'][0] > \
                    preprocessing_setup['resize'][1]:
                image_shape = (
                    preprocessing_setup['resize'][0],
                    preprocessing_setup['resize'][1] +
                    preprocessing_setup['padding'][0] +
                    preprocessing_setup['padding'][1],
                )
            else:
                image_shape = (
                    preprocessing_setup['resize'][0] +
                    preprocessing_setup['padding'][2] +
                    preprocessing_setup['padding'][3],
                    preprocessing_setup['resize'][1],
                )
        else:
            image_shape = (
                preprocessing_setup['resize'][0],
                preprocessing_setup['resize'][1],
            )
    return (image_shape, number_of_channels)


def align_image(
        template_img: np.ndarray, frame: np.ndarray, n_features: int,
        top_match_perc: float,
        uniform_scaling: bool | None = None) -> np.ndarray:
    """
    Aligns the input frame to a reference image using ORB feature matching.

    Parameters:
    - img_ref_path (str): Path to the reference image.
    - frame (np.ndarray): Input frame to be aligned.
    - n_features (int): Number of features for ORB detector.
    - top_match_perc (float): Percentage of top matches to consider.
    - blur (tuple[int, int, int] | None): Blur parameters for the images.
    - preprocessing_setup (dict[str, Any] | None): Preprocessing setup.

    Returns:
    - np.ndarray: Aligned image.
    """

    # Store the dimensions of the image
    height, width = template_img.shape[0], template_img.shape[1]

    # Create ORB DETECTOR (with BRUTE-FORCE matcher)
    orb_detector = cv2.ORB_create(n_features)

    # Find keypoints and descriptors
    kp1, d1 = orb_detector.detectAndCompute(frame, None)
    kp2, d2 = orb_detector.detectAndCompute(template_img, None)

    matcher = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)

    # Match the two sets of descriptors
    matches = matcher.match(d1, d2)

    # Convert the tuple into a list
    matches_list = list(matches)
    # Sort the list according to Hamming distance
    matches_list.sort(key=lambda x: x.distance)
    # Convert the list into a tuple
    matches = tuple(matches_list)

    # Take the top matches
    matches = matches[:int(len(matches)*top_match_perc)]
    no_of_matches = len(matches)

    # Define empty matrices of shape [no_of_matches * 2]
    p1 = np.zeros((no_of_matches, 2))
    p2 = np.zeros((no_of_matches, 2))

    # Extract the coordinates of the keypoints in the images
    for i in range(len(matches)):
        p1[i, :] = kp1[matches[i].queryIdx].pt
        p2[i, :] = kp2[matches[i].trainIdx].pt

    # Find the homography matrix
    affine_matrix, _ = cv2.estimateAffinePartial2D(p1, p2)

    if uniform_scaling:
        # Apply the affine transformation (translation and rotation only)
        transformed_img = cv2.warpAffine(frame, affine_matrix, (width, height))
    else:
        # Extract translation and rotation from affine_matrix
        translation = affine_matrix[:, 2]
        rotation = np.arctan2(affine_matrix[1, 0], affine_matrix[0, 0])
        # Build a new matrix without uniform scaling
        new_affine_matrix = np.zeros((2, 3))
        new_affine_matrix[0, 0] = np.cos(rotation)
        new_affine_matrix[0, 1] = -np.sin(rotation)
        new_affine_matrix[1, 0] = np.sin(rotation)
        new_affine_matrix[1, 1] = np.cos(rotation)
        new_affine_matrix[:, 2] = translation
        # Applica the affine transformation without scaling
        transformed_img = cv2.warpAffine(frame, new_affine_matrix,
                                         (width, height))

    return transformed_img


def resize_image(original_image: np.ndarray,
                 target_size: tuple[int, int] = (320, 320)) -> np.ndarray:
    """
    Resize the input image to the target size using padding or cropping.

    Parameters:
    - original_image (np.ndarray): Input image to be resized.
    - target_size (tuple[int, int]): Target size of the resized image.

    Returns:
    - np.ndarray: Resized image.
    """

    # Ottieni le dimensioni originali dell'immagine
    original_height, original_width = original_image.shape[:2]

    # Calcola le dimensioni del padding o del cropping
    pad_width = max(0, (target_size[0] - original_width) // 2)
    pad_height = max(0, (target_size[1] - original_height) // 2)

    # Calcola il cropping (se necessario)
    crop_width = max(0, (original_width - target_size[0]) // 2)
    crop_height = max(0, (original_height - target_size[1]) // 2)

    # Assicurati che le dimensioni siano intere
    # e distribuisci il cropping o il padding in modo equo
    pad_width_left = pad_width // 2
    pad_width_right = pad_width - pad_width_left
    pad_height_top = pad_height // 2
    pad_height_bottom = pad_height - pad_height_top

    crop_width_left = crop_width // 2
    crop_width_right = crop_width - crop_width_left
    crop_height_top = crop_height // 2
    crop_height_bottom = crop_height - crop_height_top

    # Esegui il cropping o il padding
    if original_width < target_size[0] or original_height < target_size[1]:
        # Padding
        padded_image = cv2.copyMakeBorder(original_image,
                                          pad_height_top, pad_height_bottom,
                                          pad_width_left, pad_width_right,
                                          cv2.BORDER_CONSTANT,
                                          value=[128, 128, 128])
        resized_image = padded_image
    else:
        # Cropping
        cropped_image = original_image[
            crop_height_top:original_height-crop_height_bottom,
            crop_width_left:original_width-crop_width_right]
        resized_image = cropped_image

    # Ridimensiona l'immagine alle dimensioni desiderate
    resized_image = cv2.resize(resized_image, target_size)

    return resized_image


def verify_excel_file_existence(file_path: str) -> None:
    """
    Verifies the existence of an Excel file.
    Creates a new file if it doesn't exist.

    Parameters:
    - file_path (str): Path to the Excel file.
    """
    if not os.path.isfile(file_path):
        print(f"'{file_path}' file does not exist. Creating file...")
        # Create a new Excel file with an empty sheet
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print(f"'{file_path}' file successfully created.")


def verify_excel_sheet_existence(file_path: str, sheet_name: str) -> bool:
    """
    Verifies the existence of a sheet in an Excel file.
    Creates the sheet if it doesn't exist.

    Parameters:
    - file_path (str): Path to the Excel file.
    - sheet_name (str): Name of the sheet to verify.

    Returns:
    - bool: True if the sheet already exists, False if it was created.
    """
    try:
        # Open Excel file
        workbook = openpyxl.load_workbook(file_path)
        # Verify its existence
        if sheet_name not in workbook.sheetnames:
            # If it does not exist, create it
            workbook.create_sheet(
                title=sheet_name, index=len(workbook.sheetnames))
            workbook.save(file_path)
            return False
        else:
            return True
    except FileNotFoundError:
        print(f"WARNING: '{file_path}' file does not exist.")
        return False
    except Exception as e:
        print(f"WARNING: {str(e)}")
        return False
